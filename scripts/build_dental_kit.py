from __future__ import annotations

import csv
import shutil
import textwrap
import zipfile
from pathlib import Path

from PIL import Image, ImageDraw, ImageFont, ImageFilter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Image as RLImage,
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover
    Workbook = None


REPO_ROOT = Path(__file__).resolve().parents[1]
DESKTOP_ROOT = REPO_ROOT.parent
LOCAL_ROOT = DESKTOP_ROOT / "Clinicas Dentales"
DOWNLOAD_ROOT = REPO_ROOT / "downloads" / "clinicas-dentales" / "materiales"
AFFILIATE_ROOT = REPO_ROOT / "downloads" / "affiliate-kit-dental"
WEB_ASSETS = REPO_ROOT / "assets" / "dental"
LOGO = REPO_ROOT / "logo-prontia.jpg"

IMAGE_SOURCES = {
    "clinic": Path("/Users/luiscurras/.codex/generated_images/019f4aab-e9d3-7671-b9a9-9ba46c71f5dc/call_YGrH8lAxLvtpMVPYdw2OIMhW.png"),
    "reception": Path("/Users/luiscurras/.codex/generated_images/019f4aab-e9d3-7671-b9a9-9ba46c71f5dc/call_A2FG7RyUxDUdTyzqpndPLJTP.png"),
    "consultation": Path("/Users/luiscurras/.codex/generated_images/019f4aab-e9d3-7671-b9a9-9ba46c71f5dc/call_DxxttqZhVICqhGWv7KVg82XI.png"),
}

FONT_REGULAR = Path("/System/Library/Fonts/Supplemental/Arial.ttf")
FONT_BOLD = Path("/System/Library/Fonts/Supplemental/Arial Bold.ttf")
FONT_BLACK = Path("/System/Library/Fonts/Supplemental/Arial Black.ttf")

TITLE = "Kit Agenda Dental Llena 30 Días"
PRODUCT_SLUG = "clinicas-dentales"
PRODUCT_ZIP = REPO_ROOT / "downloads" / "kit-agenda-dental-llena-30-dias.zip"
AFFILIATE_ZIP = REPO_ROOT / "downloads" / "kit-base-afiliados-dental.zip"


def reset_dir(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path)
    path.mkdir(parents=True, exist_ok=True)


def ensure_dirs() -> None:
    for path in [LOCAL_ROOT, DOWNLOAD_ROOT, AFFILIATE_ROOT, WEB_ASSETS]:
        path.mkdir(parents=True, exist_ok=True)


def font(path: Path, size: int) -> ImageFont.FreeTypeFont:
    return ImageFont.truetype(str(path), size)


def fit_image(src: Path, size: tuple[int, int]) -> Image.Image:
    image = Image.open(src).convert("RGB")
    target_w, target_h = size
    ratio = max(target_w / image.width, target_h / image.height)
    resized = image.resize((int(image.width * ratio), int(image.height * ratio)), Image.LANCZOS)
    left = (resized.width - target_w) // 2
    top = (resized.height - target_h) // 2
    return resized.crop((left, top, left + target_w, top + target_h))


def draw_wrapped(draw: ImageDraw.ImageDraw, xy: tuple[int, int], text: str, font_obj, fill, max_width: int, line_gap: int = 6) -> int:
    x, y = xy
    lines = []
    for raw_line in text.split("\n"):
        words = raw_line.split()
        current = ""
        for word in words:
            candidate = f"{current} {word}".strip()
            if draw.textbbox((0, 0), candidate, font=font_obj)[2] <= max_width:
                current = candidate
            else:
                if current:
                    lines.append(current)
                current = word
        if current:
            lines.append(current)
    for line in lines:
        draw.text((x, y), line, font=font_obj, fill=fill)
        y += draw.textbbox((0, 0), line, font=font_obj)[3] + line_gap
    return y


def paste_logo(draw_img: Image.Image, x: int, y: int, max_w: int) -> None:
    logo = Image.open(LOGO).convert("RGBA")
    ratio = max_w / logo.width
    logo = logo.resize((max_w, int(logo.height * ratio)), Image.LANCZOS)
    draw_img.paste(logo, (x, y), logo)


def rounded_rectangle(draw: ImageDraw.ImageDraw, box, radius, fill, outline=None, width=1):
    draw.rounded_rectangle(box, radius=radius, fill=fill, outline=outline, width=width)


def create_square_cover(path: Path) -> None:
    img = fit_image(IMAGE_SOURCES["clinic"], (1080, 1080)).filter(ImageFilter.GaussianBlur(radius=0.2))
    overlay = Image.new("RGBA", img.size, (0, 0, 0, 0))
    od = ImageDraw.Draw(overlay)
    od.rectangle((0, 0, 610, 1080), fill=(255, 255, 255, 232))
    od.rectangle((610, 0, 1080, 1080), fill=(4, 40, 60, 22))
    od.ellipse((-90, 865, 170, 1125), fill=(84, 19, 125, 236))
    od.ellipse((892, -110, 1185, 180), fill=(242, 148, 36, 238))
    img = Image.alpha_composite(img.convert("RGBA"), overlay)
    draw = ImageDraw.Draw(img)
    paste_logo(img, 72, 66, 74)
    draw.text((166, 88), "ProntIA LATAM", font=font(FONT_BOLD, 32), fill=(23, 27, 38))
    y = 214
    y = draw_wrapped(draw, (72, y), "KIT AGENDA\nDENTAL LLENA\n30 DÍAS", font(FONT_BLACK, 52), (80, 20, 120), 490, 0)
    y += 18
    y = draw_wrapped(
        draw,
        (72, y),
        "Para clínicas dentales con más citas, menos huecos y más tratamientos aceptados.",
        font(FONT_REGULAR, 30),
        (27, 31, 44),
        500,
        8,
    )
    y += 24
    pills = [
        "WhatsApp + agenda",
        "Seguimiento de presupuestos",
        "100 prompts de IA",
        "Reseñas, CRM y campañas",
    ]
    for pill in pills:
        rounded_rectangle(draw, (72, y, 572, y + 48), 24, (247, 234, 211, 255))
        draw.text((102, y + 11), pill, font=font(FONT_BOLD, 25), fill=(27, 31, 44))
        y += 62
    rounded_rectangle(draw, (72, 982, 642, 1038), 0, (255, 250, 242, 245))
    draw.text((102, 997), "Producto descargable en prontialatam.com", font=font(FONT_REGULAR, 24), fill=(75, 70, 82))
    path.parent.mkdir(parents=True, exist_ok=True)
    img.convert("RGB").save(path, quality=92, optimize=True)


def create_vertical_cover(path: Path) -> None:
    img = fit_image(IMAGE_SOURCES["clinic"], (1080, 1350))
    overlay = Image.new("RGBA", img.size, (2, 35, 51, 92))
    card = Image.new("RGBA", img.size, (0, 0, 0, 0))
    cd = ImageDraw.Draw(card)
    rounded_rectangle(cd, (70, 80, 720, 870), 34, (255, 255, 255, 236))
    cd.ellipse((870, -96, 1168, 190), fill=(242, 148, 36, 238))
    cd.ellipse((-118, 1118, 198, 1436), fill=(84, 19, 125, 238))
    img = Image.alpha_composite(Image.alpha_composite(img.convert("RGBA"), overlay), card)
    draw = ImageDraw.Draw(img)
    paste_logo(img, 106, 128, 72)
    draw.text((196, 149), "ProntIA LATAM", font=font(FONT_BOLD, 31), fill=(23, 27, 38))
    y = 260
    y = draw_wrapped(draw, (106, y), "KIT AGENDA\nDENTAL LLENA\n30 DÍAS", font(FONT_BLACK, 52), (80, 20, 120), 555, 0)
    y += 22
    draw_wrapped(draw, (106, y), "Más citas, menos huecos y más tratamientos aceptados.", font(FONT_REGULAR, 30), (27, 31, 44), 540, 8)
    draw.text((106, 760), "Clínicas dentales LATAM", font=font(FONT_BOLD, 31), fill=(18, 91, 120))
    path.parent.mkdir(parents=True, exist_ok=True)
    img.convert("RGB").save(path, quality=90, optimize=True)


def create_horizontal_banner(path: Path) -> None:
    img = fit_image(IMAGE_SOURCES["consultation"], (1600, 900))
    overlay = Image.new("RGBA", img.size, (0, 0, 0, 0))
    od = ImageDraw.Draw(overlay)
    od.rectangle((0, 0, 760, 900), fill=(255, 255, 255, 235))
    od.rectangle((760, 0, 1600, 900), fill=(0, 74, 94, 30))
    img = Image.alpha_composite(img.convert("RGBA"), overlay)
    draw = ImageDraw.Draw(img)
    paste_logo(img, 82, 72, 76)
    draw.text((180, 96), "ProntIA LATAM", font=font(FONT_BOLD, 34), fill=(23, 27, 38))
    y = 210
    y = draw_wrapped(draw, (82, y), "Kit Agenda Dental\nLlena 30 Días", font(FONT_BLACK, 60), (80, 20, 120), 600, 8)
    y += 20
    draw_wrapped(draw, (82, y), "IA, WhatsApp, campañas y seguimiento para clínicas dentales de LATAM.", font(FONT_REGULAR, 32), (27, 31, 44), 590, 10)
    path.parent.mkdir(parents=True, exist_ok=True)
    img.convert("RGB").save(path, quality=90, optimize=True)


def copy_images() -> dict[str, Path]:
    assets = {
        "base": WEB_ASSETS / "fondo_clinica_dental_tablet.jpg",
        "reception": WEB_ASSETS / "recepcion_whatsapp_clinica_dental.jpg",
        "consultation": WEB_ASSETS / "consulta_plan_tratamiento_dental.jpg",
        "square": WEB_ASSETS / "producto_cuadrado_kit_agenda_dental_llena.jpg",
        "vertical": WEB_ASSETS / "producto_vertical_kit_agenda_dental_llena.jpg",
        "banner": WEB_ASSETS / "header_banner_kit_agenda_dental_llena.jpg",
        "logo": WEB_ASSETS / "logo_prontia_latam.jpg",
    }
    WEB_ASSETS.mkdir(parents=True, exist_ok=True)
    for key, src_key in [("base", "clinic"), ("reception", "reception"), ("consultation", "consultation")]:
        img = Image.open(IMAGE_SOURCES[src_key]).convert("RGB")
        img.save(assets[key], quality=88, optimize=True)
    shutil.copy2(LOGO, assets["logo"])
    create_square_cover(assets["square"])
    create_vertical_cover(assets["vertical"])
    create_horizontal_banner(assets["banner"])
    return assets


def write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(textwrap.dedent(content).strip() + "\n", encoding="utf-8")


def write_csv(path: Path, rows: list[list[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(rows)


def copy_to_roots(relative: str, content: str) -> None:
    write_text(LOCAL_ROOT / relative, content)
    write_text(DOWNLOAD_ROOT / relative, content)


def copy_csv_to_roots(relative: str, rows: list[list[str]]) -> None:
    write_csv(LOCAL_ROOT / relative, rows)
    write_csv(DOWNLOAD_ROOT / relative, rows)


def build_content_files() -> None:
    readme = f"""
    # {TITLE}

    Bienvenido al kit de ProntIA LATAM para clínicas dentales.

    Empieza por el dossier principal en `01_Dossier_Principal`, después abre el calendario de 30 días y finalmente carga las plantillas de WhatsApp en el equipo de recepción.

    ## Orden recomendado

    1. Leer el diagnóstico y mapa de fuga de facturación.
    2. Configurar la hoja CRM de seguimiento.
    3. Guardar las plantillas de WhatsApp.
    4. Ejecutar la campaña de 30 días por semanas.
    5. Medir citas, presupuestos enviados, tratamientos aceptados y reseñas.

    Este material ayuda a ordenar comunicación, marketing, seguimiento y experiencia del paciente. No sustituye criterio clínico, asesoría legal ni normativa sanitaria local.
    """
    copy_to_roots("00_LEEME/LEEME_PRIMERO.md", readme)

    dossier_md = f"""
    # {TITLE}

    ## Promesa

    IA, WhatsApp, campañas y seguimiento para clínicas dentales de LATAM que quieren más citas, menos huecos y más tratamientos aceptados sin depender de improvisación diaria.

    ## Dolor principal

    La clínica puede vender tratamientos rentables, pero pierde ingresos cuando las consultas se quedan en WhatsApp, las citas no se confirman, los presupuestos no se siguen y los pacientes antiguos no vuelven a entrar en agenda.

    ## Sistema en 5 bloques

    1. Captación: campañas por tratamiento y dolor visible.
    2. Conversación: guiones de recepción para no vender solo por precio.
    3. Agenda: confirmaciones, recordatorios y recuperación de huecos.
    4. Presupuesto: seguimiento profesional de planes de tratamiento.
    5. Reputación: reseñas, referidos y reactivación.

    ## Rutina diaria de 20 minutos

    - Revisar nuevos mensajes y clasificarlos.
    - Confirmar citas de mañana.
    - Hacer seguimiento de presupuestos pendientes.
    - Reactivar dos pacientes antiguos.
    - Pedir una reseña a pacientes satisfechos.

    ## Métricas mínimas

    - Consultas recibidas.
    - Citas agendadas.
    - Citas asistidas.
    - Presupuestos enviados.
    - Presupuestos aceptados.
    - Ticket estimado.
    - Reseñas solicitadas y publicadas.
    """
    copy_to_roots("01_Dossier_Principal/Kit_Agenda_Dental_Llena_30_Dias_ProntIA_LATAM.md", dossier_md)

    calendar_rows = [["Dia", "Objetivo", "Accion", "Canal", "Mensaje base", "KPI"]]
    weekly_focus = [
        ("Ordenar recepción", "Configurar CRM, respuestas rápidas y agenda"),
        ("Captar demanda", "Publicar campañas por tratamiento y dolor"),
        ("Convertir presupuestos", "Hacer seguimiento y explicar valor"),
        ("Reactivar y fidelizar", "Reseñas, referidos y pacientes antiguos"),
    ]
    actions = [
        ("Auditar los últimos 30 WhatsApps", "WhatsApp", "Detecta preguntas repetidas y respuestas que pierden citas.", "Consultas clasificadas"),
        ("Configurar etiquetas del CRM", "Sheets/Excel", "Nuevo, citado, asistió, presupuesto, aceptado, reactivar.", "CRM preparado"),
        ("Crear respuesta de bienvenida", "WhatsApp", "Gracias por escribirnos. Para orientarte mejor, ¿qué tratamiento te interesa y en qué zona estás?", "Tiempo de respuesta"),
        ("Confirmar citas de mañana", "WhatsApp", "Te esperamos mañana a las [hora]. Si necesitas cambiar, avísanos hoy para liberar el espacio.", "Confirmaciones"),
        ("Publicar campaña de diagnóstico", "Instagram/Facebook", "¿Hace más de 6 meses que no revisas tu salud dental? Agenda una valoración.", "Mensajes recibidos"),
        ("Reactivar pacientes antiguos", "WhatsApp", "Hola [nombre], vimos que hace tiempo no vienes a control. ¿Quieres que revisemos agenda esta semana?", "Reactivados"),
        ("Pedir reseñas a pacientes satisfechos", "Google", "Tu opinión ayuda a otros pacientes a confiar en la clínica.", "Reseñas solicitadas"),
    ]
    for day in range(1, 31):
        focus = weekly_focus[min((day - 1) // 8, 3)]
        action = actions[(day - 1) % len(actions)]
        calendar_rows.append([str(day), focus[0], f"{focus[1]} - {action[0]}", action[1], action[2], action[3]])
    copy_csv_to_roots("02_Calendario_30_Dias/calendario_30_dias_clinicas_dentales.csv", calendar_rows)
    calendar_md = "\n".join([f"- Día {row[0]}: {row[2]} ({row[3]}). KPI: {row[5]}" for row in calendar_rows[1:]])
    copy_to_roots("02_Calendario_30_Dias/Calendario_30_Dias_Campanas_Clinicas_Dentales.md", f"# Calendario 30 Días\n\n{calendar_md}")

    whatsapp_rows = [["Situacion", "Objetivo", "Mensaje"]]
    whatsapp_templates = [
        ("Primer contacto", "Responder sin vender solo por precio", "Hola [nombre], gracias por escribir a [clínica]. Para orientarte bien, ¿qué tratamiento te interesa y en qué ciudad/zona estás?"),
        ("Pregunta por precio", "Llevar a valoración", "El precio depende del diagnóstico y del plan indicado por el doctor. Podemos agendar una valoración para darte una opción real y evitarte información incompleta."),
        ("Confirmación 24h", "Reducir no-show", "Hola [nombre], confirmamos tu cita mañana a las [hora]. Responde CONFIRMO para mantener tu espacio."),
        ("No confirma", "Recuperar respuesta", "Hola [nombre], necesitamos confirmar tu cita de mañana. Si no puedes asistir, avísanos para ofrecerte otro horario."),
        ("No show", "Reagendar", "Hola [nombre], hoy no pudimos verte en la clínica. ¿Quieres que te proponga dos horarios para reagendar esta semana?"),
        ("Presupuesto enviado", "Dar seguimiento", "Hola [nombre], ¿pudiste revisar el plan de tratamiento que te explicamos? Si quieres, te ayudamos a resolver dudas de tiempos, pagos o fases."),
        ("Objeción caro", "Reencuadrar valor", "Te entiendo. Lo importante es comparar diagnóstico, materiales, experiencia y seguimiento, no solo el precio inicial. Podemos revisar opciones por fases."),
        ("Financiación", "Facilitar decisión", "Podemos orientarte con alternativas de pago o dividir el tratamiento por etapas cuando el caso lo permite. Lo revisamos contigo según tu diagnóstico."),
        ("Post tratamiento", "Cuidar experiencia", "Hola [nombre], ¿cómo te has sentido después de tu cita? Cualquier molestia o duda, responde este mensaje y te orientamos."),
        ("Pedir reseña", "Mejorar reputación", "Nos alegra haberte atendido. Si tu experiencia fue buena, tu reseña en Google ayuda a otros pacientes a elegir con confianza: [link]."),
    ]
    for row in whatsapp_templates:
        whatsapp_rows.append(list(row))
    copy_csv_to_roots("03_WhatsApp_Agenda_Presupuestos/plantillas_whatsapp_clinicas_dentales.csv", whatsapp_rows)
    copy_to_roots(
        "03_WhatsApp_Agenda_Presupuestos/Plantillas_WhatsApp_Agenda_Presupuestos.md",
        "# Plantillas WhatsApp, Agenda y Presupuestos\n\n" + "\n\n".join([f"## {a}\n**Objetivo:** {b}\n\n{c}" for a, b, c in whatsapp_templates]),
    )

    categories = [
        ("Captación local", "crear una campaña local para captar pacientes interesados en [tratamiento] en [ciudad]"),
        ("WhatsApp", "mejorar una respuesta de WhatsApp para convertir una consulta en cita"),
        ("Presupuestos", "hacer seguimiento de un presupuesto dental sin sonar insistente"),
        ("Reseñas", "pedir una reseña de Google de forma elegante y breve"),
        ("RRSS", "crear un post educativo para Instagram sobre [tema dental]"),
        ("Reels", "crear un guion de Reel de 20 segundos para [tratamiento]"),
        ("CRM", "clasificar pacientes por estado comercial y próxima acción"),
        ("Reactivación", "reactivar pacientes que no visitan la clínica hace más de 6 meses"),
        ("Campañas", "diseñar una campaña mensual sin prometer resultados clínicos"),
        ("Objeciones", "responder a pacientes que dicen caro, luego veo o estoy comparando"),
    ]
    prompt_rows = [["N", "Categoria", "Prompt"]]
    prompt_md_parts = ["# 100 Prompts de IA para Clínicas Dentales"]
    n = 1
    for cat, base in categories:
        prompt_md_parts.append(f"\n## {cat}")
        for i in range(10):
            prompt = f"Actúa como experto en marketing dental LATAM y ayúdame a {base}. Usa lenguaje claro, ético, sin prometer resultados médicos, con CTA a valoración y una versión corta para WhatsApp. Caso: [describe tu clínica, ciudad, tratamiento, precio orientativo y objeción frecuente]."
            prompt_rows.append([str(n), cat, prompt])
            prompt_md_parts.append(f"{n}. {prompt}")
            n += 1
    copy_csv_to_roots("04_Prompts_IA/100_prompts_ia_clinicas_dentales.csv", prompt_rows)
    copy_to_roots("04_Prompts_IA/100_Prompts_IA_Clinicas_Dentales.md", "\n".join(prompt_md_parts))

    reels = []
    for i, treatment in enumerate(["implantes", "ortodoncia invisible", "blanqueamiento", "limpieza dental", "carillas", "prótesis", "urgencias", "niños", "periodoncia", "revisión preventiva"], start=1):
        reels.append(f"## Guion {i}: {treatment.title()}\nGancho: ¿Estás posponiendo {treatment} por falta de tiempo o dudas?\nEscena 1: recepción mostrando agenda.\nEscena 2: doctora explicando en tablet.\nEscena 3: CTA: agenda una valoración y resuelve tus dudas antes de decidir.")
    copy_to_roots("05_RRSS_Reels_Stories/30_Guiones_Reels_Shorts_Dental.md", "# Guiones Reels, Shorts y Stories\n\n" + "\n\n".join(reels * 3))

    promos = [
        ("Valoración + diagnóstico inicial", "Atraer pacientes indecisos sin regalar tratamientos complejos."),
        ("Semana de limpieza preventiva", "Llenar huecos con servicios de baja fricción."),
        ("Plan de ortodoncia por fases", "Reducir objeción de desembolso inicial."),
        ("Campaña sonrisa antes de eventos", "Activar blanqueamiento y estética dental con fecha concreta."),
        ("Referidos familiares", "Incentivar recomendaciones sin prometer descuentos agresivos."),
        ("Reactivación de pacientes dormidos", "Volver a citar controles pendientes."),
        ("Urgencias con cupos limitados reales", "Organizar demanda sin saturar agenda."),
        ("Pack revisión familiar", "Captar varios pacientes de un mismo hogar."),
        ("Seguimiento de implantes pendientes", "Recuperar presupuestos de alto valor."),
        ("Carillas con evaluación previa", "Filtrar pacientes con intención real."),
    ]
    promo_text = "\n\n".join([f"## {name}\nObjetivo: {goal}\nMensaje base: Agenda una valoración en [clínica] y recibe una orientación clara sobre opciones, fases y próximos pasos." for name, goal in promos])
    copy_to_roots("06_Campanas_Bonos_Financiacion/20_Campanas_Bonos_Dentales.md", "# Campañas, Bonos y Financiación Responsable\n\n" + promo_text)

    reviews = """
    # Kit de Reseñas y Reputación Dental

    ## Cuándo pedir la reseña

    Pídela después de una experiencia positiva: final de tratamiento, control satisfactorio o solución de una urgencia.

    ## Mensaje breve

    Hola [nombre], gracias por confiar en [clínica]. Si tu experiencia fue buena, tu reseña en Google nos ayuda mucho y orienta a otros pacientes de la zona: [link].

    ## Respuesta a reseña positiva

    Gracias, [nombre]. Nos alegra saber que te sentiste bien atendido/a. Seguiremos trabajando para ofrecer una experiencia cercana, clara y profesional.

    ## Respuesta a reseña negativa

    Hola [nombre], sentimos que tu experiencia no haya sido la esperada. Nos gustaría revisar lo ocurrido con cuidado. Por favor escríbenos a [contacto] para poder ayudarte de forma directa.
    """
    copy_to_roots("07_Resenas_Reputacion/Kit_Resenas_Reputacion_Dental.md", reviews)

    scripts = """
    # Scripts de Recepción y Cierre de Tratamientos

    ## Cuando el paciente pide precio

    Entiendo que quieras saber el precio antes de venir. En odontología, el valor depende del diagnóstico, materiales, complejidad y fases del tratamiento. Lo más responsable es verte primero y explicarte opciones reales.

    ## Cuando dice "lo tengo que pensar"

    Claro, es una decisión importante. ¿Qué parte te gustaría revisar mejor: precio, tiempo, molestias, financiación o prioridad del tratamiento?

    ## Cuando compara con otra clínica

    Es normal comparar. Te recomiendo comparar diagnóstico, experiencia, materiales, seguimiento y garantías explicadas por escrito, no solo el número final.

    ## Cuando no responde al presupuesto

    Hola [nombre], retomo tu plan de tratamiento para saber si te quedó alguna duda. Si quieres, podemos ayudarte a revisar fases o tiempos para que decidas con más claridad.
    """
    copy_to_roots("11_Scripts_Recepcion_Cierre/Scripts_Recepcion_Cierre_Tratamientos.md", scripts)

    video = """
    # Prompts para Reel / Short del Kit Agenda Dental

    ## Dolor principal

    Tu clínica no pierde dinero por falta de tratamientos rentables: lo pierde en WhatsApps sin seguimiento, citas vacías y presupuestos que nadie vuelve a tocar.

    ## Guion 5 segundos

    Plano 1: recepción mirando la agenda con huecos.
    Plano 2: doctora revisando tablet.
    Plano 3: texto en pantalla: Agenda Dental Llena 30 Días.
    Voz: Más citas, menos huecos y más tratamientos aceptados con IA, WhatsApp y seguimiento.
    """
    copy_to_roots("10_Video_Reel_Short/README_VIDEO_REEL_SHORT.md", video)

    brief = """
    # Brief Canva / Web

    Producto: Kit Agenda Dental Llena 30 Días.
    Claim: Más citas, menos huecos y más tratamientos aceptados.
    Visual: clínica dental moderna, doctora latina con tablet, tonos blanco, teal, azul y naranja ProntIA.
    Evitar: promesas médicas, sangre, procedimientos invasivos, descuentos agresivos.
    """
    copy_to_roots("09_Creatividades_Web_Canva/BRIEF_WEB_CANVA_Kit_Agenda_Dental.md", brief)


def build_crm() -> None:
    headers = [
        "Fecha",
        "Paciente",
        "Teléfono",
        "Tratamiento de interés",
        "Origen",
        "Estado",
        "Valor estimado",
        "Próxima acción",
        "Fecha seguimiento",
        "Responsable",
        "Notas",
    ]
    rows = [
        ["2026-08-03", "Ejemplo Paciente", "+52 ...", "Implantes", "WhatsApp", "Presupuesto enviado", "1200", "Resolver financiación", "2026-08-05", "Recepción", "Interesado, compara opciones"],
    ]
    for root in [LOCAL_ROOT, DOWNLOAD_ROOT]:
        csv_path = root / "08_CRM_KPIs" / "crm_kpis_clinicas_dentales.csv"
        write_csv(csv_path, [headers] + rows)
        if Workbook:
            wb = Workbook()
            ws = wb.active
            ws.title = "CRM Dental"
            ws.append(headers)
            for row in rows:
                ws.append(row)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="12385B")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            widths = [13, 22, 16, 24, 18, 22, 16, 28, 18, 18, 36]
            for idx, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = width
            ws.freeze_panes = "A2"
            path = root / "08_CRM_KPIs" / "Agenda_Dental_CRM_KPIs.xlsx"
            path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(path)


def copy_creatives(assets: dict[str, Path]) -> None:
    for root in [LOCAL_ROOT, DOWNLOAD_ROOT]:
        target = root / "09_Creatividades_Web_Canva"
        target.mkdir(parents=True, exist_ok=True)
        for name, src in [
            ("producto_cuadrado_kit_agenda_dental_llena.jpg", assets["square"]),
            ("producto_vertical_kit_agenda_dental_llena.jpg", assets["vertical"]),
            ("header_banner_kit_agenda_dental_llena.jpg", assets["banner"]),
            ("fondo_clinica_dental_tablet.jpg", assets["base"]),
            ("recepcion_whatsapp_clinica_dental.jpg", assets["reception"]),
            ("consulta_plan_tratamiento_dental.jpg", assets["consultation"]),
            ("logo_prontia_latam.jpg", assets["logo"]),
        ]:
            shutil.copy2(src, target / name)


def make_pdf(assets: dict[str, Path]) -> None:
    for root in [LOCAL_ROOT, DOWNLOAD_ROOT]:
        output = root / "01_Dossier_Principal" / "Kit_Agenda_Dental_Llena_30_Dias_ProntIA_LATAM.pdf"
        output.parent.mkdir(parents=True, exist_ok=True)
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name="CoverTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=34, leading=36, textColor=colors.HexColor("#4F1478"), alignment=TA_LEFT, spaceAfter=14))
        styles.add(ParagraphStyle(name="Sub", parent=styles["BodyText"], fontName="Helvetica", fontSize=13, leading=18, textColor=colors.HexColor("#293344"), spaceAfter=10))
        styles.add(ParagraphStyle(name="H1P", parent=styles["Heading1"], fontName="Helvetica-Bold", fontSize=19, leading=23, textColor=colors.HexColor("#12385B"), spaceBefore=10, spaceAfter=8))
        styles.add(ParagraphStyle(name="H2P", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=14, leading=18, textColor=colors.HexColor("#185A78"), spaceBefore=8, spaceAfter=6))
        styles.add(ParagraphStyle(name="BodyP", parent=styles["BodyText"], fontName="Helvetica", fontSize=10.5, leading=15, textColor=colors.HexColor("#2C3440"), spaceAfter=7))
        styles.add(ParagraphStyle(name="SmallP", parent=styles["BodyText"], fontName="Helvetica", fontSize=8.5, leading=11, textColor=colors.HexColor("#5E6673"), spaceAfter=5))
        styles.add(ParagraphStyle(name="Callout", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=12, leading=16, textColor=colors.HexColor("#12385B"), backColor=colors.HexColor("#F4F8F8"), borderColor=colors.HexColor("#C9DADD"), borderWidth=0.5, borderPadding=8, spaceAfter=10))

        def on_page(canvas, doc):
            canvas.saveState()
            canvas.drawImage(str(LOGO), 0.58 * inch, 10.16 * inch, width=0.45 * inch, height=0.45 * inch, preserveAspectRatio=True, mask="auto")
            canvas.setFont("Helvetica-Bold", 8.5)
            canvas.setFillColor(colors.HexColor("#12385B"))
            canvas.drawString(1.1 * inch, 10.32 * inch, "ProntIA LATAM | Kit Agenda Dental Llena 30 Días")
            canvas.setStrokeColor(colors.HexColor("#DCE8EA"))
            canvas.line(0.58 * inch, 10.08 * inch, 7.92 * inch, 10.08 * inch)
            canvas.setFont("Helvetica", 8)
            canvas.setFillColor(colors.HexColor("#6C7785"))
            canvas.drawRightString(7.92 * inch, 0.45 * inch, f"Página {doc.page}")
            canvas.restoreState()

        doc = SimpleDocTemplate(str(output), pagesize=letter, leftMargin=0.62 * inch, rightMargin=0.62 * inch, topMargin=0.86 * inch, bottomMargin=0.72 * inch)
        story = []
        story.append(RLImage(str(assets["banner"]), width=7.25 * inch, height=4.08 * inch))
        story.append(Spacer(1, 14))
        story.append(Paragraph(TITLE, styles["CoverTitle"]))
        story.append(Paragraph("IA, WhatsApp, campañas y seguimiento para clínicas dentales de LATAM que quieren más citas, menos huecos y más tratamientos aceptados.", styles["Sub"]))
        story.append(Paragraph("El objetivo no es que el doctor se convierta en marketer. El objetivo es que recepción tenga un sistema simple para responder, agendar, confirmar, seguir presupuestos y reactivar pacientes con ayuda de IA.", styles["Callout"]))
        story.append(PageBreak())

        sections = [
            ("1. Diagnóstico del dolor", [
                "Las clínicas dentales suelen tener tratamientos rentables, pero la facturación se escapa en los puntos menos clínicos: mensajes sin responder a tiempo, pacientes que preguntan precio y desaparecen, citas sin confirmar, presupuestos pendientes y reseñas no solicitadas.",
                "Este kit convierte esas fugas en una rutina comercial diaria para recepción y dirección, sin exigir conocimientos técnicos de IA."
            ]),
            ("2. Mapa de fuga de facturación", [
                "Consulta nueva: si no se responde con estructura, el paciente compara por precio.",
                "Cita agendada: si no se confirma, aparece el hueco de última hora.",
                "Diagnóstico: si no se explica valor, el paciente aplaza.",
                "Presupuesto: si no hay seguimiento, el tratamiento queda dormido.",
                "Post-atención: si no se pide reseña, se pierde reputación pública."
            ]),
            ("3. Sistema Agenda Dental Llena", [
                "Captar demanda con campañas por tratamiento.",
                "Convertir WhatsApps en citas mediante preguntas y guiones.",
                "Reducir no-shows con confirmaciones y recordatorios.",
                "Aumentar aceptación con seguimiento de presupuestos.",
                "Reactivar pacientes antiguos y pedir reseñas."
            ]),
        ]
        for title, items in sections:
            story.append(Paragraph(title, styles["H1P"]))
            for item in items:
                story.append(Paragraph(item, styles["BodyP"]))

        story.append(RLImage(str(assets["reception"]), width=7.25 * inch, height=3.81 * inch))
        story.append(Paragraph("Recepción debe trabajar con respuestas rápidas, estados del paciente y próxima acción visible.", styles["SmallP"]))

        story.append(Paragraph("4. Plan de 30 días por semanas", styles["H1P"]))
        data = [
            ["Semana", "Objetivo", "Acciones clave"],
            ["1", "Ordenar recepción", "CRM, etiquetas, respuestas rápidas, confirmaciones y auditoría de WhatsApp."],
            ["2", "Captar demanda", "Campañas por implantes, ortodoncia, limpieza, blanqueamiento y revisión."],
            ["3", "Convertir presupuestos", "Seguimiento, objeciones, financiación y tratamiento por fases."],
            ["4", "Reactivar y fidelizar", "Pacientes antiguos, reseñas, referidos y agenda de control."],
        ]
        table = Table(data, colWidths=[0.8 * inch, 1.75 * inch, 4.55 * inch])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#12385B")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("LEADING", (0, 0), (-1, -1), 12),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D6E2E5")),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#FBFCFC")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ]))
        story.append(table)
        story.append(Spacer(1, 12))

        story.append(Paragraph("5. Guiones esenciales de WhatsApp", styles["H1P"]))
        for title, objective, msg in [
            ("Precio", "Evitar comparación superficial", "El precio depende del diagnóstico. Te podemos agendar una valoración para darte una opción real y explicarte fases."),
            ("Confirmación", "Reducir no-show", "Te esperamos mañana a las [hora]. Responde CONFIRMO para mantener tu espacio."),
            ("Cambio de cita", "Proteger agenda", "Claro, podemos ayudarte a reagendar. Te propongo [hora 1] o [hora 2]. Si liberas tu espacio hoy, podemos ofrecerlo a otro paciente."),
            ("Paciente frío", "Recuperar conversación", "Hola [nombre], retomo tu consulta para saber si sigues interesado/a en [tratamiento]. ¿Quieres que revisemos agenda esta semana?"),
            ("Presupuesto", "Reactivar decisión", "¿Pudiste revisar el plan? Podemos ayudarte con dudas de tiempo, pagos o prioridad del tratamiento."),
            ("Tratamiento por fases", "Facilitar aceptación", "Si te preocupa hacerlo todo de una vez, podemos explicarte si tu caso permite trabajar por etapas y qué conviene priorizar."),
            ("Paciente antiguo", "Reactivar control", "Hola [nombre], hace tiempo no te vemos en control. ¿Quieres que te proponga horarios para una revisión preventiva?"),
            ("Reseña", "Reputación", "Si tu experiencia fue buena, tu reseña ayuda a otros pacientes a elegir con confianza."),
        ]:
            story.append(KeepTogether([Paragraph(title, styles["H2P"]), Paragraph(f"<b>Objetivo:</b> {objective}<br/><b>Mensaje:</b> {msg}", styles["BodyP"])]))

        story.append(Paragraph("Matriz rápida de estado del paciente", styles["H1P"]))
        state_data = [
            ["Estado", "Qué significa", "Próxima acción"],
            ["Nuevo lead", "Preguntó por WhatsApp o redes", "Responder, clasificar tratamiento y proponer valoración."],
            ["Citado", "Tiene hora asignada", "Confirmar 24 horas antes y registrar canal de origen."],
            ["Presupuesto enviado", "Ya conoce plan o rango", "Resolver objeciones, fases y forma de pago."],
            ["Dormido", "No responde o aplaza", "Reactivar con mensaje corto y dos opciones de agenda."],
            ["Paciente satisfecho", "Terminó atención positiva", "Pedir reseña, referidos y programar control."],
        ]
        t_state = Table(state_data, colWidths=[1.25 * inch, 2.15 * inch, 3.5 * inch])
        t_state.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#185A78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D6E2E5")),
            ("FONTSIZE", (0, 0), (-1, -1), 8.6),
            ("LEADING", (0, 0), (-1, -1), 11),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(t_state)

        story.append(PageBreak())
        story.append(RLImage(str(assets["consultation"]), width=7.25 * inch, height=4.83 * inch))
        story.append(Paragraph("La aceptación de tratamiento mejora cuando el paciente entiende opciones, fases y próximo paso.", styles["SmallP"]))
        story.append(Paragraph("6. Campañas por tratamiento", styles["H1P"]))
        for treatment in ["Implantes", "Ortodoncia", "Blanqueamiento", "Limpieza preventiva", "Carillas", "Prótesis", "Urgencias", "Revisión familiar"]:
            story.append(Paragraph(f"<b>{treatment}:</b> publicar dolor visible, explicar valor, ofrecer valoración y cerrar con WhatsApp. Evitar prometer resultados clínicos o descuentos imposibles.", styles["BodyP"]))

        story.append(Paragraph("7. Rutina diaria de recepción", styles["H1P"]))
        for item in [
            "Responder nuevos mensajes con plantilla base y pregunta de clasificación.",
            "Confirmar citas de mañana antes de mediodía.",
            "Registrar cada paciente en CRM con estado y próxima acción.",
            "Hacer seguimiento de 3 presupuestos pendientes.",
            "Pedir 1 reseña al día cuando la experiencia haya sido positiva.",
        ]:
            story.append(Paragraph(f"• {item}", styles["BodyP"]))

        story.append(Paragraph("8. Uso de IA sin conocimientos técnicos", styles["H1P"]))
        story.append(Paragraph("La auxiliar solo necesita copiar el prompt, completar los campos entre corchetes y revisar que el texto sea correcto para la clínica. La IA no diagnostica, no promete resultados y no sustituye al doctor.", styles["BodyP"]))

        story.append(Paragraph("9. Checklist de implementación", styles["H1P"]))
        checklist = [["Acción", "Responsable", "Estado"], ["CRM creado", "Dirección/recepción", ""], ["Respuestas rápidas guardadas", "Recepción", ""], ["Calendario de 30 días asignado", "Marketing/recepción", ""], ["Link de reseñas preparado", "Recepción", ""], ["Seguimiento de presupuestos activo", "Doctor/recepción", ""]]
        t2 = Table(checklist, colWidths=[3.2 * inch, 2.3 * inch, 1.4 * inch])
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F1478")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D6E2E5")),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        story.append(t2)
        story.append(Spacer(1, 10))
        story.append(Paragraph("Aviso: adapta mensajes, precios, promociones y procesos a la regulación sanitaria, publicitaria y de protección de datos de tu país.", styles["SmallP"]))
        doc.build(story, onFirstPage=on_page, onLaterPages=on_page)


def build_affiliate_kit(assets: dict[str, Path]) -> None:
    reset_dir(AFFILIATE_ROOT)
    docs = AFFILIATE_ROOT / "docs"
    social = AFFILIATE_ROOT / "social"
    brand = AFFILIATE_ROOT / "brand"
    planning = AFFILIATE_ROOT / "planning"
    for p in [docs, social, brand, planning]:
        p.mkdir(parents=True, exist_ok=True)
    write_text(docs / "00_LEEME_PRIMERO.txt", "Kit base de afiliados para promocionar el Kit Agenda Dental Llena 30 Días. Sustituye [TU ENLACE] por tu enlace personal del portal.")
    write_text(docs / "02_DOSSIER_PRODUCTO_DENTAL.txt", "Producto: Kit Agenda Dental Llena 30 Días.\nDolor: clínicas dentales pierden citas, presupuestos y pacientes por falta de sistema comercial.\nPromesa: IA, WhatsApp, campañas y seguimiento para más citas, menos huecos y más tratamientos aceptados.\nPrecio recomendado: 47 USD pago único.")
    write_text(docs / "03_COPIES_Y_GANCHOS_DENTAL.txt", "Gancho 1: Tu clínica no pierde dinero en el sillón: lo pierde en WhatsApps sin seguimiento.\nGancho 2: Citas vacías, presupuestos dormidos y pacientes antiguos sin volver.\nCTA: Descubre el Kit Agenda Dental Llena 30 Días en [TU ENLACE].")
    write_text(docs / "04_FAQ_Y_OBJECIONES_DENTAL.txt", "Objeción: No sé usar IA.\nRespuesta: El kit trae prompts y plantillas listas para copiar y adaptar.\n\nObjeción: Mi clínica ya tiene agenda.\nRespuesta: Entonces el foco es reducir huecos, mejorar seguimiento y recuperar presupuestos.")
    write_text(docs / "05_PLAN_7_DIAS_PUBLICACION.txt", "Día 1 dolor agenda.\nDía 2 WhatsApp.\nDía 3 presupuestos.\nDía 4 reseñas.\nDía 5 campaña por tratamiento.\nDía 6 caso de uso.\nDía 7 oferta con CTA.")
    write_text(docs / "06_REGLAS_COMISIONES_Y_USO.txt", "No prometas resultados clínicos, ingresos garantizados ni tratamientos asegurados. Habla de organización, seguimiento, marketing y experiencia del paciente. Comisión base: 60% sobre venta neta cuando aplique.")
    write_csv(planning / "Calendario_7_dias_dental.csv", [["Dia", "Tema", "CTA"], ["1", "Agenda con huecos", "[TU ENLACE]"], ["2", "WhatsApp que no convierte", "[TU ENLACE]"], ["3", "Presupuestos sin seguimiento", "[TU ENLACE]"], ["4", "Reseñas y confianza", "[TU ENLACE]"], ["5", "Implantes/ortodoncia", "[TU ENLACE]"], ["6", "Recepción con IA", "[TU ENLACE]"], ["7", "Oferta Kit Dental", "[TU ENLACE]"]])
    for name, src in [
        ("logo_prontia_latam.jpg", assets["logo"]),
        ("producto_cuadrado_kit_agenda_dental_llena.jpg", assets["square"]),
        ("producto_vertical_kit_agenda_dental_llena.jpg", assets["vertical"]),
        ("header_banner_kit_agenda_dental_llena.jpg", assets["banner"]),
    ]:
        shutil.copy2(src, brand / name)
        shutil.copy2(src, social / name)


def zip_dir(source: Path, zip_path: Path, arc_prefix: str | None = None) -> None:
    if zip_path.exists():
        zip_path.unlink()
    zip_path.parent.mkdir(parents=True, exist_ok=True)
    zip_resolved = zip_path.resolve()
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for file in sorted(source.rglob("*")):
            if file.is_file():
                if file.resolve() == zip_resolved:
                    continue
                rel = file.relative_to(source)
                arcname = Path(arc_prefix) / rel if arc_prefix else rel
                zf.write(file, arcname.as_posix())


def main() -> None:
    for src in IMAGE_SOURCES.values():
        if not src.exists():
            raise FileNotFoundError(src)
    if not LOGO.exists():
        raise FileNotFoundError(LOGO)
    ensure_dirs()
    reset_dir(LOCAL_ROOT)
    reset_dir(DOWNLOAD_ROOT)
    assets = copy_images()
    build_content_files()
    build_crm()
    copy_creatives(assets)
    make_pdf(assets)
    build_affiliate_kit(assets)
    zip_dir(LOCAL_ROOT, LOCAL_ROOT / "99_Entrega_Web_ZIP" / "Kit_Agenda_Dental_Llena_30_Dias_ProntIA_LATAM.zip", "Clinicas Dentales")
    (DOWNLOAD_ROOT / "99_Entrega_Web_ZIP").mkdir(parents=True, exist_ok=True)
    shutil.copy2(LOCAL_ROOT / "99_Entrega_Web_ZIP" / "Kit_Agenda_Dental_Llena_30_Dias_ProntIA_LATAM.zip", DOWNLOAD_ROOT / "99_Entrega_Web_ZIP" / "Kit_Agenda_Dental_Llena_30_Dias_ProntIA_LATAM.zip")
    shutil.copy2(LOCAL_ROOT / "99_Entrega_Web_ZIP" / "Kit_Agenda_Dental_Llena_30_Dias_ProntIA_LATAM.zip", PRODUCT_ZIP)
    zip_dir(AFFILIATE_ROOT, AFFILIATE_ZIP, "affiliate-kit-dental")
    print(f"Kit local: {LOCAL_ROOT}")
    print(f"Kit web: {DOWNLOAD_ROOT}")
    print(f"Producto ZIP: {PRODUCT_ZIP}")
    print(f"Afiliados ZIP: {AFFILIATE_ZIP}")
    print(f"Portada: {assets['square']}")


if __name__ == "__main__":
    main()
